"""Builds an in-memory .xlsx workbook (one sheet per topic) from NormalizedView instances."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import NormalizedView

_SHEET_TITLES = {
    'technical': 'Technical Compliance',
    'price': 'Price Compliance',
}


def _write_sheet(workbook: Workbook, title: str, view: NormalizedView) -> None:
    sheet = workbook.create_sheet(title=title[:31])  # Excel sheet name limit
    headers = [*view.tender_columns, *view.bid_columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDDDDD')

    for row in view.rows:
        tender_values = [row.tender_cells.get(col, '') for col in view.tender_columns]
        bid_cell_values = [row.bid_values.get(col) or '' for col in view.bid_columns]
        sheet.append([*tender_values, *bid_cell_values])

    for col_index in range(1, len(headers) + 1):
        is_tender_column = col_index <= len(view.tender_columns)
        sheet.column_dimensions[get_column_letter(col_index)].width = 40 if is_tender_column else 30


def build_workbook(views: list[NormalizedView]) -> bytes:
    """Build a workbook with one sheet per view (Technical Compliance / Price Compliance)."""
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default empty sheet
    for view in views:
        title = _SHEET_TITLES.get(view.topic.value, view.topic.value)
        _write_sheet(workbook, title, view)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
